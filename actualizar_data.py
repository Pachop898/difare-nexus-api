"""
actualizar_data.py — ETL Difare Nexus

Regenera api/data.db leyendo TODOS los Excels de la carpeta excels/.
Autodetecta tipo de archivo por la primera columna:
  - 'FECHA' (YYYYMM) -> archivo mensual oficial -> tabla `ventas`
  - 'DIA'   (diario) -> archivo semanal SAP      -> tabla `sap`

Uso:
    python actualizar_data.py

Flujo semanal:
    1. Pega el nuevo Excel en excels/
    2. python actualizar_data.py
    3. git add api/data.db && git commit -m "..." && git push

Cuando llegue el mensual oficial (dia 12 aprox):
    1. Borra los semanales del mes cerrado
    2. Pega el Reporte_Mensual en excels/
    3. python actualizar_data.py
    4. git push

La regla de precedencia (mensual > semanal por mes) la aplica la app en
tiempo de query: si un mes existe en `ventas`, los dias de ese mes en `sap`
se ignoran automaticamente para esa farmacia.
"""

import os
import sys
import sqlite3
import openpyxl
from glob import glob

ROOT = os.path.dirname(os.path.abspath(__file__))
EXCELS = os.path.join(ROOT, "excels")
DB = os.path.join(ROOT, "api", "data.db")

COLS = [
    "FECHA", "CANAL", "UNIDAD", "SEGMENTO", "GRUPOCLIENTE",
    "IDPROPIETARIO", "PROPIETARIO", "IDESTABLECIMIENTO", "ESTABLECIMIENTO",
    "RUC", "GRUPOPDV", "CODIGOPDV", "POS", "IDCORPORACION", "CORPORACION",
    "PROVEEDOR", "IDNEPTUNO", "IDDIFARE", "MARCA", "PRODUCTO",
    "REGION", "PROVINCIA", "CIUDAD", "UNIDADES_ROTADAS",
    "VENTA NETA RECUPERO", "STOCK", "STOCK_VALORIZADO", "DIA"
]

DDL_VENTAS = """
CREATE TABLE ventas (
    FECHA INTEGER, CANAL TEXT, UNIDAD TEXT, SEGMENTO TEXT, GRUPOCLIENTE TEXT,
    IDPROPIETARIO REAL, PROPIETARIO TEXT, IDESTABLECIMIENTO TEXT, ESTABLECIMIENTO TEXT,
    RUC TEXT, GRUPOPDV TEXT, CODIGOPDV TEXT, POS TEXT, IDCORPORACION INTEGER, CORPORACION TEXT,
    PROVEEDOR TEXT, IDNEPTUNO INTEGER, IDDIFARE INTEGER, MARCA TEXT, PRODUCTO TEXT,
    REGION TEXT, PROVINCIA TEXT, CIUDAD TEXT, UNIDADES_ROTADAS REAL,
    "VENTA NETA RECUPERO" REAL, STOCK REAL, STOCK_VALORIZADO REAL, DIA TEXT
)
"""

DDL_SAP = """
CREATE TABLE sap (
    DIA TEXT, CANAL TEXT, UNIDAD TEXT, SEGMENTO TEXT, GRUPOCLIENTE TEXT,
    IDPROPIETARIO REAL, PROPIETARIO TEXT, IDESTABLECIMIENTO TEXT, ESTABLECIMIENTO TEXT,
    RUC TEXT, GRUPOPDV TEXT, CODIGOPDV TEXT, POS TEXT, IDCORPORACION INTEGER, CORPORACION TEXT,
    PROVEEDOR TEXT, IDNEPTUNO INTEGER, IDDIFARE INTEGER, MARCA TEXT, PRODUCTO TEXT,
    REGION TEXT, PROVINCIA TEXT, CIUDAD TEXT, UNIDADES_ROTADAS REAL,
    "VENTA NETA RECUPERO" REAL, STOCK REAL, STOCK_VALORIZADO REAL
)
"""

INSERT_VENTAS = """
INSERT INTO ventas (FECHA, CANAL, UNIDAD, SEGMENTO, GRUPOCLIENTE, IDPROPIETARIO, PROPIETARIO,
    IDESTABLECIMIENTO, ESTABLECIMIENTO, RUC, GRUPOPDV, CODIGOPDV, POS, IDCORPORACION, CORPORACION,
    PROVEEDOR, IDNEPTUNO, IDDIFARE, MARCA, PRODUCTO, REGION, PROVINCIA, CIUDAD, UNIDADES_ROTADAS,
    "VENTA NETA RECUPERO", STOCK, STOCK_VALORIZADO, DIA)
VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
"""

INSERT_SAP = """
INSERT INTO sap (DIA, CANAL, UNIDAD, SEGMENTO, GRUPOCLIENTE, IDPROPIETARIO, PROPIETARIO,
    IDESTABLECIMIENTO, ESTABLECIMIENTO, RUC, GRUPOPDV, CODIGOPDV, POS, IDCORPORACION, CORPORACION,
    PROVEEDOR, IDNEPTUNO, IDDIFARE, MARCA, PRODUCTO, REGION, PROVINCIA, CIUDAD, UNIDADES_ROTADAS,
    "VENTA NETA RECUPERO", STOCK, STOCK_VALORIZADO)
VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
"""


def detect_type(header):
    """'FECHA' como primer col -> mensual (ventas); 'DIA' -> semanal (sap)."""
    first = (header[0] or "").strip().upper() if header else ""
    if first == "FECHA":
        return "ventas"
    if first == "DIA":
        return "sap"
    return None


def clean(v):
    if isinstance(v, str):
        return v.strip() or None
    return v


def load_file(path, conn):
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR abriendo: {e}")
        return 0, None

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return 0, None

    tipo = detect_type(header)
    if not tipo:
        print(f"  SKIP (formato desconocido, primera col: {header[0]!r})")
        return 0, None

    if len(header) < 27:
        print(f"  SKIP (columnas insuficientes: {len(header)})")
        return 0, None

    cur = conn.cursor()
    batch = []
    n = 0
    for row in rows_iter:
        if row is None or all(x is None for x in row):
            continue
        r = [clean(v) for v in row[:27]]
        while len(r) < 27:
            r.append(None)
        if tipo == "ventas":
            # FECHA (YYYYMM) tambien se guarda en DIA como string
            fecha = r[0]
            dia_str = str(fecha) if fecha is not None else None
            batch.append(tuple(r) + (dia_str,))
        else:
            batch.append(tuple(r))
        if len(batch) >= 5000:
            cur.executemany(INSERT_VENTAS if tipo == "ventas" else INSERT_SAP, batch)
            n += len(batch)
            batch.clear()
    if batch:
        cur.executemany(INSERT_VENTAS if tipo == "ventas" else INSERT_SAP, batch)
        n += len(batch)
    conn.commit()
    wb.close()
    return n, tipo


def main():
    if not os.path.isdir(EXCELS):
        print(f"No existe carpeta {EXCELS}")
        sys.exit(1)

    files = sorted(glob(os.path.join(EXCELS, "*.xlsx")))
    if not files:
        print(f"No hay Excels en {EXCELS}")
        sys.exit(1)

    print(f"Regenerando {DB}")
    os.makedirs(os.path.dirname(DB), exist_ok=True)
    if os.path.exists(DB):
        os.remove(DB)
    conn = sqlite3.connect(DB)
    conn.execute(DDL_VENTAS)
    conn.execute(DDL_SAP)
    conn.commit()

    tot_v, tot_s = 0, 0
    for f in files:
        name = os.path.basename(f)
        print(f"  · {name}")
        n, tipo = load_file(f, conn)
        if tipo == "ventas":
            tot_v += n
            print(f"    -> ventas: {n:,} filas")
        elif tipo == "sap":
            tot_s += n
            print(f"    -> sap: {n:,} filas")

    # Indices utiles
    print("Creando indices...")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ventas_pos ON ventas(POS)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ventas_unidad ON ventas(UNIDAD)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ventas_fecha ON ventas(FECHA)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sap_pos ON sap(POS)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sap_unidad ON sap(UNIDAD)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sap_dia ON sap(DIA)")
    conn.commit()

    # Resumen
    v = conn.execute("SELECT COUNT(*) FROM ventas").fetchone()[0]
    s = conn.execute("SELECT COUNT(*) FROM sap").fetchone()[0]
    meses_v = [r[0] for r in conn.execute(
        "SELECT DISTINCT substr(DIA,1,6) FROM ventas ORDER BY 1").fetchall()]
    meses_s = [r[0] for r in conn.execute(
        "SELECT DISTINCT CASE WHEN INSTR(DIA,'/')>0 "
        "THEN substr(DIA,1,4)||substr(DIA,6,2) ELSE substr(DIA,1,6) END "
        "FROM sap ORDER BY 1").fetchall()]

    # Meses en conflicto (presentes en ambas -> la app dara precedencia a ventas)
    solape = sorted(set(meses_v) & set(meses_s))

    conn.close()
    size_mb = os.path.getsize(DB) / (1024 * 1024)
    print()
    print("=" * 50)
    print(f"OK  data.db: {size_mb:.1f} MB")
    print(f"    ventas: {v:,} filas | meses: {meses_v}")
    print(f"    sap:    {s:,} filas | meses: {meses_s}")
    if solape:
        print(f"    Solape mensual>semanal (ignorado en app): {solape}")
    print("=" * 50)
    print()
    print("Siguiente paso:")
    print('  git add api/data.db && git commit -m "Data actualizada" && git push')


if __name__ == "__main__":
    main()
